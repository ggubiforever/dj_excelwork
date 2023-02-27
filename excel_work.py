import sys
from PyQt5 import uic
import openpyxl
from openpyxl import load_workbook
from PyQt5.QtWidgets import *
import re
#import win32com.client
import pandas as pd
import os
import xlsxwriter


form_class = uic.loadUiType("main_windows.ui")[0]

class MyWindow(QMainWindow, form_class):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.cond = 0
        self.bom_data = pd.DataFrame()
        self.df_bom = pd.DataFrame()

    def openfile1(self):  #엑셀데이타 합치기
        cond = self.cond
        if cond == 0:
            QMessageBox.about(self, "Warning!!!", "Bom 먼저 읽어라")
            return
        path = os.getcwd()
        fileName = QFileDialog.getOpenFileNames(self, "파일선택", path, "Excel Files (*.xls *.xlsx)")
        if fileName:
            self.merge_excel(fileName)
        else:
            QMessageBox.about(self, "notice", "선택된 파일이 없소")

    def openfile2(self):
        exopt = "2"
        path = os.getcwd()
        fileName = QFileDialog.getOpenFileName(self, "파일선택", path, "Excel Files (*.xls *.xlsx)")
        if fileName:
            self.getexcelfile(fileName)
        else:
            QMessageBox.about(self, "notice", "선택된 파일이 없소")

    def openfile3(self):  #엑셀데이타 합치기
        path = os.getcwd()
        fileName = QFileDialog.getOpenFileNames(self, "파일선택", path, "Excel Files (*.xls *.xlsx)")
        if fileName:
            self.merge_excel(fileName)
        else:
            QMessageBox.about(self, "notice", "선택된 파일이 없소")

    def openfile4(self): #경진 BOM 읽어오기
        QMessageBox.about(self, "notice", "원본 엑셀파일 첫번째 행은 항목명이 되어야 함")
        path = os.getcwd()
        fileName = QFileDialog.getOpenFileName(self,"파일선택",path, "Excel Files (*.xls, *.xlsx)")
        if fileName:
            self.mk_bom_kj(fileName)
        else:
            QMessageBox.about(self,"notice","선택된 파일 없다.")

    def openfile5(self): #동진섬유 수입잔량 확인
        path = os.getcwd()
        fileName = QFileDialog.getOpenFileName(self,"파일선택",path,"Excel Files (*.xls *.xlsx)")
        self.make_simple_impData(fileName)

    def openfile6(self): ## kyungjin export of heungsin making BOM code
        path = os.getcwd()
        fileName = QFileDialog.getOpenFileName(self,"파일선택",path,"Excel Files (*.xls *.xlsx)")
        self.make_kyungjin_heungsin_Bom_code(fileName)


    def getexcelfile(self,fileName): #BOM 저장
        fname = fileName
        bom_data = pd.DataFrame()
        df = pd.read_excel(fname[0], sheet_name = "정리")
        bom_data = bom_data.append(df, ignore_index=True)
        bom_data = self.bom_process(bom_data)
        self.bom_data = bom_data

    def make_simple_impData(self,fname):
        path = os.getcwd()
        df_org = pd.read_excel(fname[0])
        df_f = pd.read_excel(path + "\\"+'수입원장_간단서식.xlsx')
        org_lst = ['신고번호', '란번호', '수리일자', '세번부호', '자재코드', '자재(환급)물량', '물량단위', '결제금액', '관세액', '계산잔량(사용가능물량)', '계산잔량세액(사용가능세액)', '신고금액','규격번호']
        fin_lst = ['수입신고번호', '란번호', '수입수리일', '세번부호', '자재코드', '물량', '물량단위', '규격결제금액', '규격별관세', '잔량물량', '잔량관세',
                   '규격별CIF원화','규격번호']
        df_f[fin_lst] = df_org[org_lst]
        df_org['란결제금액'] = df_org.groupby(['신고번호','란번호'])['결제금액'].transform('sum')
        df_org['란신고금액'] = df_org.groupby(['신고번호','란번호'])['신고금액'].transform('sum')
        df_org['란관세'] = df_org.groupby(['신고번호','란번호'])['관세액'].transform('sum')
        df_f[['결제금액','과세가격(원화)','수입관세']] = df_org[['란결제금액','란신고금액','란관세']]
        df_f[['통화단위','규격통화단위']] = 'USD'
        df_f[['원재료구분']] = '00'
        df_f['란번호'] = df_f['란번호'].astype(str)
        df_f['규격번호'] = df_f['규격번호'].astype(str)
        df_f['란번호'] = df_f['란번호'].str.zfill(3)
        df_f['규격번호'] = df_f['규격번호'].str.zfill(2)
        df_f.fillna("",inplace = True)
        df_f.to_excel('c:\\temp\\impo_simpleform.xlsx')
        os.startfile('c:\\temp\\impo_simpleform.xlsx')
        QMessageBox.about(self,"notice","수입원장엑셀간단서식완료")


    def bom_process(self,df):

        for df_idx, df_row in df.iterrows():
            gt = "제품DJTNO"
            gt2 = "제품명[△1]"
            df_text = df_row[gt]
            df_text2 = df_row[gt2]
            regex_text = r'DJT?\s*-?\w+'
            # regex_text = r'DJT?\s*-?\w+-[A-Z0-9]*'
            result = self.reg_logic1(regex_text,df_text)
            df.loc[df_idx, "분류코드"] = result
            result2 = self.reg_logic2(df_text2)
            g_text = ''
            for i in range(len(result2)):
                gk_restext = str(result2[i])
                gk_restext = gk_restext.replace(' ', '')
                if g_text == '':
                    g_text = g_text + gk_restext
                else:
                    g_text = g_text + "," + gk_restext
            df.loc[df_idx, "구분"] = g_text

        self.cond = 1
        QMessageBox.about(self, "notice", "BOM 읽기 완료")
        df_bom = df.astype('str')
        lst_pd = ['분류코드', '구분']
        df_bom['key'] = df_bom[lst_pd].apply(lambda x: ' '.join(x), axis=1)  # pandas 텍스트를 서로 합쳐서 값을 찾을거다
        self.df_bom = df_bom
        return df

    def mk_bom_kj(self,fileName):
        df = pd.read_excel(fileName[0])
        df_jepum = pd.DataFrame(columns=['제품코드','세번코드','품명','규격','물량단위','소요량산정방법','규격2','적용기간1','적용기간2','양수자용제품식별번호'])
        df_jaje = pd.DataFrame(columns = ['제품코드','자재코드','자재-세번부호','자재품명','자재-규격','자재-소요식','자재-단위소요량','자재-물량단위','조사란구분','자재-규격2','동일자재여부'])
        df_jaje[["제품코드","자재코드","자재-단위소요량"]] = df[["제품코드","자재코드","소요량"]]
        df_jaje = df_jaje.fillna("")
        df_jaje['자재-물량단위'] = 'KG'
        df_jepum[['제품코드','품명','물량단위']] = df[['제품코드','제품품명','단위']]
        df_jepum['소요량산정방법'] = '02'
        df_jepum.fillna("")
        df_jepum = df_jepum.drop_duplicates(['제품코드'],keep='first')

        df3 = pd.DataFrame(columns=['제품코드','자재코드','대체자재코드','대체자재-세번부호','대체자재-품명','대체자재-규격','대체자재-소요식','대체자재-단위소요량','대체자재-물량단위','조사란구분','소요량 환산식','소요량 환산율'
])
        df4 = pd.DataFrame(columns=['제품코드','자재코드','부산물코드','부산물-세번부호','부산물-품명','부산물-물량단위','부산물-발생율','공제구분','공제율','부산물가격','공제계산'
])
        df5 = pd.DataFrame(columns=['제품코드','자재코드','대체자재코드','부산물코드','부산물-세번부호','부산물-품명','부산물-물량단위','부산물-발생율','공제구분','공제율','부산물가격'
])
        writer = pd.ExcelWriter('c:\\temp\\kj_bom_ncom.xlsx', engine='xlsxwriter')
        df_jepum.to_excel(writer, sheet_name='제품사항', index=False)
        df_jaje.to_excel(writer, sheet_name='제품_자재사항', index=False)
        df3.to_excel(writer, sheet_name = '제품_대체자재사항', index = False)
        df4.to_excel(writer, sheet_name='제품_부산물사항', index=False)
        df5.to_excel(writer, sheet_name='제품_대체자재의부산물', index=False)
        writer.save()
        QMessageBox.about(self, "작업완료", "c:\\temp 경로에 저장되었습니다.")
        os.startfile('c:\\temp')


    def vlookup_bom(self):
        return

    def make_kyungjin_heungsin_Bom_code(self,fname):
        df = pd.read_excel(fname[0],sheet_name = "모델규격사항")
        df1 = pd.read_excel(fname[0], sheet_name = "공통사항")
        df2 = pd.read_excel(fname[0], sheet_name="란사항")
        df4 = pd.read_excel(fname[0], sheet_name = "모델규격사항_서류첨부")
        df['품명2'].fillna("A",inplace = True)
        df['품명3'].fillna("A",inplace = True)
        df['품명4'].fillna("A",inplace = True)
        df['품명5'].fillna("A", inplace=True)
        df['품명6'].fillna("A", inplace=True)
        df['품명7'].fillna("A", inplace=True)
        df['품명8'].fillna("A", inplace=True)
        df['품명키'] = df['품명1'] + df['품명2'] + df['품명3'] + df['품명4'] + df['품명5'] + df['품명6'] + df['품명7'] + df['품명8']
        regex = r'\(\d+\w*-\d+\w*\)'
        regex2 = r"\([MW]+'+[S:]+\d+\w*-\d+\w*\)"
        df['품명키'].head(100)
        p = re.compile(regex)
        p2 = re.compile(regex2)
        for idx, row in df.iterrows():
            txt = str(df.loc[idx, '품명키'])
            m = p.search(txt)
            m2 = p2.search(txt)
            t_text = df.loc[idx, '제품코드']
            if m:
                fx = m.group()
                df.loc[idx, '품명키2'] = fx
            elif m2:
                fx = m2.group()
                ctxt = "(" + fx[1] + ")" + "(" + fx[5:]
                df.loc[idx, '품명키2'] = ctxt
            else:
                df.loc[idx, '품명키2'] = ''

        df['제품코드'] = df['제품코드'].apply(lambda x: re.sub(r'(\([MW]\))*\(\d+\w*-\d+\w*\)', '', x))
        df['제품코드'] = df['제품코드'] + df['품명키2']
        df.drop('품명키', axis = 1, inplace = True)
        df.drop('품명키2', axis = 1, inplace = True)
        df3 = df
        #print(df['제품코드'])
        cond = 'K'
        self.make_out_completelyExcel(df1, df2, df3, df4, cond)


    def merge_excel(self,fileName):
        fname = fileName
        all_data1 = pd.DataFrame()
        all_data2 = pd.DataFrame()
        all_data3 = pd.DataFrame()
        for cnt in range(len(fname[0])):
            df1 = pd.read_excel(fname[0][cnt],sheet_name="공통사항")
            df2 = pd.read_excel(fname[0][cnt], sheet_name="란사항")
            df3 = pd.read_excel(fname[0][cnt],sheet_name="모델규격사항")
           # df3 = pd.merge(df3,df2[['신고번호','세번부호']],on='신고번호', how='left')
            # df3 = pd.merge(df3,df1[['신고번호','수리일']],on='신고번호', how='left')
            all_data3 = all_data3.append(df3, ignore_index=True)
            all_data2 = all_data2.append(df2, ignore_index=True)
            all_data1 = all_data1.append(df1, ignore_index=True)

        self.all_data1 = all_data1
        self.all_data2 = all_data2
        self.all_data3 = all_data3
        all_data4 = pd.DataFrame()
        whajusangho = all_data1.loc[2,'제조자통관고유부호']
        if whajusangho == '경진섬유1171011' or whajusangho == '중앙섬유1011018':
            self.make_out_completelyExcel(all_data1,all_data2,all_data3,all_data4,cond=0)
        else:
            all_data3 = self.insert_code()
            self.all_data3 = all_data3
            self.cond = 0
            QMessageBox.about(self, "notice", "동진섬유 읽기완료")

    def reg_logic1(self,regex_text,df_text):
        reg_re = re.compile(regex_text)
        result = reg_re.search(str(df_text))
        if result:
            result = result.group()
            t_result = re.sub(r'\s+', "-", str(result))
            t_result = re.sub(r'DJ',"DJ-", t_result)
            t_result = re.sub(r'DJ-T',"DJT", t_result)
            t_result = re.sub(r'DJT', "DJT-",t_result)
            t_result = re.sub(r'--',"-",t_result)
            return t_result

    def reg_logic2(self,df_text):
        regex_txt = r'EPM\s*[1-9]|EPM\s*[1-9*][(]AD[)]|CDP-P|P-CDP|CDP|\sP\s|\sP$|N-CDP|CDP-N|P-N|N-P|CP-P|C-P|WPN|DTY'
        reg_re = re.compile(regex_txt)
        result = reg_re.findall(str(df_text))
        print(result)
        return result

#    def getcodekyungjin(self,df):


    def getdataFromBom(self,text,gbn): ## 여기서 BOM 에있는 정확한 코드를 찿을거다
        cond = self.cond
        df_bom = self.df_bom
        df_bom = df_bom[df_bom.분류코드 != '']
       # print(df_bom)
        ntext = text + ' ' + gbn
        filtered_index = df_bom[df_bom['key'] == ntext].index.tolist()
        if filtered_index:
            get_text = df_bom.loc[filtered_index[0],'제품DJTNO']
        else:
            get_text = text
        return get_text

    def insert_code(self):
        df = self.all_data3
        cond = self.cond
        for df_idx, df_row in df.iterrows():
            for p_name in range(0,9):
                if p_name == 0:
                    gt = '제품코드'
                else:
                    gt = "품명" + str(p_name)
                df_text = df_row[str(gt)]
                self.df_text = df_text
                regex_text_s = r'DJT?\s*-?\d+'
                regex_text = r'DJT?\s*-?\d+-[A-Z0-9()]*' # 2022-06-30  정규식 수정 이전 : r'DJT?\s*-?\d+-[A-Z0-9]*'
                reg_re_s = re.compile(regex_text_s)
                reg_re = re.compile(regex_text)
                reschk = reg_re_s.search(str(df_text)) #품명에 해당되는 셀의 텍스트가 reg_re_s 정규식에 매칭되면
                result = self.reg_logic1(regex_text_s,df_text)
                t_result = self.reg_logic1(regex_text,df_text)
                gk_result = self.reg_logic2(df_text)
                g_text = ''
                for i in range(len(gk_result)):
                    gk_restext = str(gk_result[i])
                    gk_restext = gk_restext.replace(' ','')
                    if g_text == '':
                        g_text = g_text + gk_restext
                    else:
                        g_text = g_text + "," + gk_restext
                if reschk: #품명에 해당되는 셀의 텍스트가 reg_re_s 정규식에 매칭되면
                    print("첫결과"+result)
                    df.loc[df_idx, "DJT원시코드"] = result #원시코드상태일때
                    df.loc[df_idx, "구분"] = g_text
                    if t_result: #원시코드에 버전이 부여된상태일때
                        df.loc[df_idx, "DJT코드(버전포함)"] = t_result
                        break
                    else:
                        get_jcode = self.getdataFromBom(result,g_text)
                        df.loc[df_idx, "DJT코드(버전포함)"] = get_jcode
                        # df.loc[df_idx, "DJT코드(버전포함)"] = result

                """reg_m = r'DJT?\s*-?\w+-[A-Z0-9]+'
                reg_rem = re.compile(reg_m)
                r_match = reg_rem.search(str(df_text))"""
                """if r_match:
                    r_text = r_match.group()
                    print("두번째결과" + r_text)
                    df.loc[df_idx,"NEW제품코드1"] = r_text
                    df.loc[df_idx, "NEW제품코드2"] = t_result
                    df.loc[df_idx, "구분"] = gk_result"""

        print(df)
        return df



    def writeexcel(self):
        cond = self.cond
        print(cond)
        if cond == 0:
            all_data1 = self.all_data1
            all_data2 = self.all_data2
            all_data3 = self.all_data3
            all_data4 = pd.DataFrame()
            self.make_out_completelyExcel(all_data1,all_data2,all_data3,all_data4,cond = 0)
            #all_data1.to_excel('result1.xlsx')
            #all_data2.to_excel('result2.xlsx')
            #all_data3.to_excel('result3.xlsx')
        else:
            all_data = self.bom_data
            print(all_data)
            all_data.to_excel('bomxls.xlsx')
        QMessageBox.about(self, "notice", "엑셀생성완료")

    def make_simpleform(self):
        self.cond = 1
        path = "C:\\Users\\jinwoo\\Documents\\동진섬유"
        fileName = QFileDialog.getOpenFileName(self, "파일선택", path, "Excel Files (*.xlsx)")
        path = os.getcwd()
        if fileName:
            self.make_simpleform_step2(fileName)
        else:
            QMessageBox.about(self, "notice", "선택된 파일이 없소")

    def make_simpleform_step2(self,fileName):
        wb_read = openpyxl.Workbook()
        sheet1 = wb_read.active
        #sheet1.title = 'Sheet1'
        titles = ['수출신고번호',	'근거서류번호',	'수출형태',	'수리일자',	'',	'',
                  '란번호',	'세번부호',	'란FOB금액',	'',	'제품코드',	'품명',	'규격',	'수출물량',	'물량단위',
                  '규격 결제금액',	'통화단위',	'제조자코드',	'수출국',	'규격번호',	'4세대여부[Y,N]',
                  '규격FOB',	'환급사용여부']
        for cnt in range(len(titles)):
            sheet1.cell(row=1, column=cnt + 1).value = titles[cnt]

        wb_open = load_workbook(fileName[0])
        ws = wb_open['수출통관현황']
        cnt_chk = "T"
        tccnt = [1,4,7,8,9,20,12,14,15,16]
        # tccnt : 1-신고번호,2-근거번호,3-수출형태,4-수리일자,5-공란,6-공란,7-란번호,8-세번부호,9-란신고금액,10-공란,11-제품코드,12-품명
        # 13-규격,14-물량,15-물량단위,16-규격결재금액,17-통화단위,18-제조자코드,19-수출국,20-규격번호
        chk_ccnt = [2,4,94,95,101,123,124,134,135,137]
        # chk_ccnt :2-신고번호,4-수리일자,94-란번호,95-세번,101-란FOB,123-규격번호,124-품명,134-수량,135-단위,137-금액,138-제품코드
        i = 0
        for ccnt in range(138):
            rcnt = 3  # 서흥 서식이 3번째 행부터 시작
            iccnt = int(ccnt) + 1
            if i == 10:
                break
            if iccnt not in chk_ccnt:
                continue
            while cnt_chk:
                print("iccnt",iccnt)
                value = ws.cell(row = rcnt, column = iccnt).value
                if value:
                    if i == 0:
                        value = str(value).replace("-", "")
                    if i == 3:
                        result = str(value).replace("-", "")
                        value = result.replace(".", "")
                    if i ==1:
                        value = str(value).replace("/","")
                    regex_text = r'DJT?\s*-?\w+-?[A-Z0-9]*'
                    reg_re = re.compile(regex_text)
                    jepumcd = reg_re.search(str(value))
                    if jepumcd:
                        value_s = jepumcd.group()
                        value_j = re.sub(r'\s+', "-", str(value_s))
                        value_j = re.sub(r'DJ', "DJ-", value_j)
                        value_j = re.sub(r'DJ-T', "DJT", value_j)
                        value_j = re.sub(r'DJT', "DJT-", value_j)
                        value_j = re.sub(r'--', "-", value_j)
                        sheet1.cell(row=int(rcnt) - 1, column = 11).value = value_j
                        sheet1.cell(row=int(rcnt) - 1, column=int(tccnt[i])).value = value
                        rcnt = rcnt + 1
                    else:
                        print("I",i)
                        sheet1.cell(row = int(rcnt) - 1, column = int(tccnt[i])).value = value
                        rcnt = rcnt + 1
                else:
                    i = i + 1
                    break
        self.save_simpleform(wb_read)

    def save_simpleform(self,wb):
        wb.save('simpleform.xlsx')

    def valuefix(self,value):
        result = value.replace("-","")
        result = result.replace(".","")

        return result

    def make_out_completelyExcel(self,f1,f2,f3,f4,cond):
        if cond == 'K':
            wfname = 'final_kyung_heungsin.xlsx'
        else:
            wfname = 'final.xlsx'
        writer = pd.ExcelWriter(wfname,engine = 'xlsxwriter')
        df2 = self.re_getnolanhng(f2)
        df3 = self.re_getnolanhng(f3)
        f1.to_excel(writer,sheet_name = '공통사항', index = False)
        df2.to_excel(writer,sheet_name = '란사항', index = False)
        df3.to_excel(writer,sheet_name = '모델규격사항', index = False)
        if f4.any:
            f4.to_excel(writer, sheet_name='모델규격사항_서류첨부', index=False)
        else:
            f4 = pd.DataFrame()
            f4.to_excel(writer,sheet_name = '모델규격사항_서류첨부',index = False)
        writer.save()
        message_t = "소스위치에" + wfname + "이름으로 저장되었습니다."
        QMessageBox.about(self, "notice", message_t)

    def re_getnolanhng(self,df):
        if df.columns.values[7] == '행번호': #모델규격사항 시트라면
            for df_idx, df_row in df.iterrows():
                lan_no = df.loc[df_idx,"란번호"]
                hng_no = df.loc[df_idx,"행번호"]
                lan_no = str(lan_no).zfill(3)
                hng_no = str(hng_no).zfill(2)
                df.loc[df_idx, "란번호"] = lan_no
                df.loc[df_idx, '행번호'] = hng_no
        else:
            for df_idx, df_row in df.iterrows():
                lan_no = df.loc[df_idx,"란번호"]
                lan_no = str(lan_no).zfill(3)
                df.loc[df_idx,"란번호"] = lan_no

        return df


    def jepum_excel_ncom(self):
        path = 'c:/'
        fileName = QFileDialog.getOpenFileNames(self, "파일선택", path, "Excel Files (*.xls *.xlsx)")
        if fileName:
            self.jepum_excel_ncom2(fileName)

    def jepum_excel_ncom2(self,fname):
        df = pd.read_excel(fname[0][0])
        df_temp1 = pd.DataFrame(columns = ['제품코드','자재코드','대체자재코드','대체자재-세번부호','대체자재-품명','대체자재-규격','대체자재-소요식','대체자재-단위소요량','대체자재-물량단위','조사란구분','소요량 환산식','소요량 환산율'])
        df_temp2 = pd.DataFrame(columns = ['제품코드',	'자재코드',	'부산물코드',	'부산물-세번부호',	'부산물-품명',	'부산물-물량단위',	'부산물-발생율',	'공제구분',	'공제율',	'부산물가격',	'공제계산'])
        df_temp3 = pd.DataFrame(columns = ['제품코드',	'자재코드',	'대체자재코드',	'부산물코드',	'부산물-세번부호',	'부산물-품명',	'부산물-물량단위',	'부산물-발생율',	'공제구분',	'공제율',	'부산물가격'])

        for i in range(2):
            df_jepum = self.make_jepum1(df,i)
            df_soyo = self.make_jepum2(df, i)
        df_jepum.to_excel('c:/temp/제품코드.xlsx',sheet_name = '제품사항')
        df_soyo.to_excel('c:/temp/제품코드.xlsx',sheet_name = '제품-자재사항')
        df_temp1.to_excel('c:/temp/제품코드.xlsx',sheet_name = '제품-대체자재사항')
        df_temp2.to_excel('c:/temp/제품코드.xlsx', sheet_name='제품-부산물사항')
        df_temp3.to_excel('c:/temp/제품코드.xlsx', sheet_name='제품_대체자재의부산물')

    def make_jepum1(self,df, i):
        tdf = pd.DataFrame(columns = ['제품코드','세번코드','품명','규격','물량단위','소요량산정방식','규격2','적용기간1','적용기간2','양수자용제품식별번호'])

        tdf['품명'] = df['제품']
        tdf['소요량산정방식'] = '02'
        if i == 0: # 단위 YD
            tdf['물량단위'] = 'YD'
            tdf['제품코드'] = df['제품DJTNO']
        else:
            tdf['물량단위'] = 'M'
            tdf['제품코드'] = tdf['제품코드'].apply(lambda x: tdf['제품코드'] + '(M)')
        tdf.drop_duplicates(['제품코드'],keep = 'first', inplace = True)
        return tdf


    def make_jepum2(self,df,i):
        tdf = pd.DataFrame(columns=['제품코드','자재코드','자재-세번부호','자재-품명','자재-규격','자재-소요식','자재-단위소요량','자재-물량단위','조사란구분','자재-규격2','동일자재여부'])
        tdf[['자재코드','자재=픔명']] = df[['원사코드','원사명(표준)']]
        tdf['자재-물량단위'] = 'KG'
        if i == 0:
            tdf['제품코드'] = df['제품DJTNO']
            tdf['자재-소요식'] = df['소요량1']
            tdf['자재-단위소요량'] = df['소요량1']
        else:
            tdf['제품코드'] = tdf['제품코드'].apply(lambda x: tdf['제품코드'] + '(M)')
            tdf['자재-소요식'] = df['소요량2']
            tdf['자재-단위소요량'] = df['소요량2']

        tdf.drop_duplicates(['제품코드','자재코드'], keep='first', inplace = True)
        return tdf









if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()